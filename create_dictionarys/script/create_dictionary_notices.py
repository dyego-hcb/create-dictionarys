import csv
import sys
import os
import pandas as pd

from openpyxl import Workbook
from unidecode import unidecode

current_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.abspath(os.path.join(current_dir, '..'))
sys.path.append(project_root)

from utils.extract_info_notices.script.extract_info_notices import extract_data
from utils.words_tokenize.script.words_tokenize import tokenize_words
from utils.words_lowercase.script.words_lowercase import convert_words_lower_case
from utils.remove_accentuation.script.remove_accentuation import remover_accentuation
from utils.remove_punctuation.script.remove_punctuation import remover_ponctuation
from utils.remove_stopwords.script.remove_stopwords import remove_stopwords_in_list
from utils.words_stemmer.script.words_stemmer import stemmize_words

def add_notice_on_dict(dict_notice, id_dict_notice, id_notice, title_notice, content_notice, classe_notice):
    words_without_ponctuation = remover_ponctuation(content_notice)
    words_without_accentuation = remover_accentuation(
        words_without_ponctuation)
    words_tokenized = tokenize_words(words_without_accentuation)
    words_lower_case = convert_words_lower_case(words_tokenized)
    words_without_stopwords = remove_stopwords_in_list(words_lower_case)
    words_stemmed = stemmize_words(words_without_stopwords)

    dict_notice[id_dict_notice] = {}
    dict_notice[id_dict_notice]['id_notice'] = id_notice
    dict_notice[id_dict_notice]['title_notice'] = title_notice
    dict_notice[id_dict_notice]['notice_content'] = words_lower_case
    dict_notice[id_dict_notice]['notice_content_without_stopwords'] = words_without_stopwords
    dict_notice[id_dict_notice]['notice_content_stemm_without_stopwords'] = words_stemmed
    dict_notice[id_dict_notice]['classe_notice'] = classe_notice
    dict_notice[id_dict_notice]['notice_words_total_with_stopwords'] = len(
        words_lower_case)
    dict_notice[id_dict_notice]['notice_words_total_without_stopwords'] = len(
        words_without_stopwords)

    return dict_notice


def create_dictionary_notices(path_bd, dict_notice, classe_notice):
    data_list = extract_data(path_bd)

    print('Starting create dict of notices ...')
    id_dict = len(dict_notice)
    words_total_with_stopwords_group = 0
    words_total_without_stopwords_group = 0

    for data in data_list[1:]:
        id_notice = data[0]
        title = data[1]
        content = data[2]
        classe = data[3]

        if classe == classe_notice:
            id_dict += 1
            dict_notice = add_notice_on_dict(
                dict_notice, id_dict, id_notice, title, content, classe)
            words_total_with_stopwords_group += dict_notice[id_dict]['notice_words_total_with_stopwords']
            words_total_without_stopwords_group += dict_notice[id_dict]['notice_words_total_without_stopwords']

    print('Finish create dict of notices\n')

    return dict_notice, words_total_with_stopwords_group, words_total_without_stopwords_group


def load_dict_notices_xlsx(folder_path, file_name, dict_notice):
    print('Starting loading dict of notices ...')
    file_path = os.path.join(folder_path, file_name)

    df = pd.read_excel(file_path)

    for row_id, row in df.iterrows():
        id_dict_notice = row_id
        id_notice = row['id_notice'] if not pd.isna(row['id_notice']) else ''
        title_notice = row['title_notice']
        notice_content = row['notice_content']
        notice_content_without_stopwords = row['notice_content_without_stopwords']
        notice_content_stemm_without_stopwords = row['notice_content_stemm_without_stopwords']
        classe_notice = row['classe_notice']
        notice_words_total_with_stopwords = row['notice_words_total_with_stopwords'] if not pd.isna(
            row['notice_words_total_with_stopwords']) else ''
        notice_words_total_without_stopwords = row['notice_words_total_without_stopwords'] if not pd.isna(
            row['notice_words_total_without_stopwords']) else ''

        print("id_dict_notice: " + str(id_dict_notice))
        print("id_notice: " + str(id_notice))
        print("title_notice: " + str(title_notice))
        print("notice_content: " + str(notice_content))
        print("notice_content_without_stopwords: " +
              str(notice_content_without_stopwords))
        print("notice_content_stemm_without_stopwords: " +
              str(notice_content_stemm_without_stopwords))
        print("classe_notice: " + str(classe_notice))
        print("notice_words_total_with_stopwords: " +
              str(notice_words_total_with_stopwords))
        print("notice_words_total_without_stopwords: " +
              str(notice_words_total_without_stopwords))

        dict_notice[id_dict_notice] = {}
        dict_notice[id_dict_notice]['id_notice'] = id_notice
        dict_notice[id_dict_notice]['title_notice'] = title_notice
        dict_notice[id_dict_notice]['notice_content'] = notice_content
        dict_notice[id_dict_notice]['notice_content_without_stopwords'] = notice_content_without_stopwords
        dict_notice[id_dict_notice]['notice_content_stemm_without_stopwords'] = notice_content_stemm_without_stopwords
        dict_notice[id_dict_notice]['classe_notice'] = classe_notice
        dict_notice[id_dict_notice]['notice_words_total_with_stopwords'] = notice_words_total_with_stopwords
        dict_notice[id_dict_notice]['notice_words_total_without_stopwords'] = notice_words_total_without_stopwords

    print('Finish load dict of notices\n')

    return dict_notice


def save_dict_notices_to_xlsx(file_path, dict_notice, group_name, words_with_stopwords, words_without_stopwords):
    print('Starting save dict of notices ...')

    df = pd.DataFrame.from_dict(dict_notice, orient='index')

    with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
        df.to_excel(
            writer, sheet_name=f'Dicionario de Noticias {group_name}', index=False)

        worksheet = writer.sheets[f'Dicionario de Noticias {group_name}']

        (max_row, max_col) = df.shape

        for col_num, value in enumerate(df.columns.values):
            worksheet.write(0, col_num, value)

        for row_num, row_data in enumerate(df.itertuples(index=False)):
            for col_num, value in enumerate(row_data):
                if isinstance(value, list):
                    df.at[row_num, df.columns[col_num]] = '\n'.join(value)
                    value = '\n'.join(value)
                worksheet.write(row_num, col_num, value)

        last_col = len(df.columns)
        df.insert(last_col, 'info_dict', '')

        info_dict_list = [
            f"- Dicionario de Noticias {group_name}",
            f"  Esse dicionario possui {len(dict_notice)} noticias",
            f"  Esse dicionario possui {words_without_stopwords} palavras (STOP-WORDS REMOVIDAS)",
            f"  Esse dicionario possui {words_with_stopwords} palavras (STOP-WORDS NO TEXTO)",
            "  Estrutura do dicionario:",
            "  ID - ID NOTICE - TITLE NOTICE - NOTICE ALL WORDS - CLASSE NOTICE - NOTICE WORDS TOTAL WITH STOPWORDS -  NOTICE WORDS TOTAL WITHOUT STOPWORDS"
        ]

        for i in range(len(info_dict_list)):
            df.at[i, 'info_dict'] = info_dict_list[i]

        df.to_excel(
            writer, sheet_name=f'Dicionario de Noticias {group_name}', index=False)

    print('Finish save dict of notices\n')


def save_dict_notices_to_csv(file_path, dict_notice, group_name, words_with_stopwords, words_without_stopwords):
    print('Starting save dict of notices ...')

    pd.set_option('display.max_colwidth', None)

    df = pd.DataFrame.from_dict(dict_notice, orient='index')

    df['title_notice'] = df['title_notice'].apply(unidecode)

    df = df.apply(lambda col: col.apply(lambda x: '\n'.join(
        map(str, x)) if isinstance(x, list) else x))

    df['info_dict'] = ''

    info_dict_list = [
        f"- Dicionario de Noticias {group_name}",
        f"  Esse dicionario possui {len(dict_notice)} noticias",
        f"  Esse dicionario possui {words_without_stopwords} palavras (STOP-WORDS REMOVIDAS)",
        f"  Esse dicionario possui {words_with_stopwords} palavras (STOP-WORDS NO TEXTO)",
        "  Estrutura do dicionario:",
        "  ID - ID NOTICE - TITLE NOTICE - NOTICE ALL WORDS - CLASSE NOTICE - NOTICE WORDS TOTAL WITH STOPWORDS -  NOTICE WORDS TOTAL WITHOUT STOPWORDS"
    ]

    df.at[1, 'info_dict'] = info_dict_list[0]
    for i in range(1, len(info_dict_list)):
        df.at[i+1, 'info_dict'] = info_dict_list[i]

    df.to_csv(file_path, index=False, encoding='utf-8',
              quoting=csv.QUOTE_NONNUMERIC)

    print('Finish save dict of notices\n')


def create_dictionary_notices_relevant_info(dict_notice_relevant_info, dict_notice):
    print('Starting create dict of notices with relevant info ...')

    for id_dict_notice in dict_notice:
        id_dict_info = len(dict_notice_relevant_info)
        notice_info = dict_notice[id_dict_notice]
        dict_notice_relevant_info[id_dict_info] = {}
        dict_notice_relevant_info[id_dict_info]['id_notice'] = notice_info.get(
            'id_notice')
        dict_notice_relevant_info[id_dict_info]['title_notice'] = notice_info.get(
            'title_notice')
        dict_notice_relevant_info[id_dict_info]['real_words_strongs_in_notice'] = [
        ]
        dict_notice_relevant_info[id_dict_info]['real_words_strongs_in_notice_number_appear'] = [
        ]
        dict_notice_relevant_info[id_dict_info]['fake_words_strongs_in_notice'] = [
        ]
        dict_notice_relevant_info[id_dict_info]['fake_words_strongs_in_notice_number_appear'] = [
        ]
        dict_notice_relevant_info[id_dict_info]['notice_strong_words_real_total'] = 0
        dict_notice_relevant_info[id_dict_info]['notice_strong_words_fake_total'] = 0
        dict_notice_relevant_info[id_dict_info]['notice_strong_words_total'] = 0
        dict_notice_relevant_info[id_dict_info]['classe_notice'] = notice_info.get(
            'classe_notice')
        dict_notice_relevant_info[id_dict_info]['notice_words_total_with_stopwords'] = notice_info.get(
            'notice_words_total_with_stopwords')
        dict_notice_relevant_info[id_dict_info]['notice_words_total_without_stopwords'] = notice_info.get(
            'notice_words_total_without_stopwords')

    print('Finish create dict of notices with relevant info\n')

    return dict_notice_relevant_info


def update_dictionary_notices_relevant_info(dict_notice_relevant_info, dict_notice, dict_strong_word, class_strong_word):
    print('Starting update dict of notices with relevant info ...')

    for id_dict_noitice_relevant_info, notice_relevant_info in dict_notice_relevant_info.items():
        real_words_strongs_in_notice = list(
            notice_relevant_info['real_words_strongs_in_notice'])
        fake_words_strongs_in_notice = list(
            notice_relevant_info['fake_words_strongs_in_notice'])
        real_words_strongs_in_notice_number_appear = notice_relevant_info[
            'real_words_strongs_in_notice_number_appear']
        fake_words_strongs_in_notice_number_appear = notice_relevant_info[
            'fake_words_strongs_in_notice_number_appear']

        id_notice_relevant_info = notice_relevant_info['id_notice']

        matching_notice = None
        for notice in dict_notice.values():
            if notice['id_notice'] == id_notice_relevant_info:
                matching_notice = notice
                break

        if matching_notice:
            notice_words = matching_notice['notice_content_stemm_without_stopwords']

            words_strongs_in_notice = []
            word_appears = {}

            for word in notice_words:
                if word in {entry['word'] for entry in dict_strong_word.values()}:
                    words_strongs_in_notice.append(word)
                    word_appears[word] = notice_words.count(word)

            if class_strong_word == 1:
                real_words_strongs_in_notice.extend(words_strongs_in_notice)
                real_words_strongs_in_notice_number_appear.extend(
                    [word_appears.get(word, 0) for word in words_strongs_in_notice])
                dict_notice_relevant_info[id_dict_noitice_relevant_info]['notice_strong_words_real_total'] += len(
                    real_words_strongs_in_notice)
                dict_notice_relevant_info[id_dict_noitice_relevant_info]['notice_strong_words_total'] += dict_notice_relevant_info[id_dict_noitice_relevant_info]['notice_strong_words_real_total']
            elif class_strong_word == 0:
                fake_words_strongs_in_notice.extend(words_strongs_in_notice)
                fake_words_strongs_in_notice_number_appear.extend(
                    [word_appears.get(word, 0) for word in words_strongs_in_notice])
                dict_notice_relevant_info[id_dict_noitice_relevant_info]['notice_strong_words_fake_total'] += len(
                    fake_words_strongs_in_notice)
                dict_notice_relevant_info[id_dict_noitice_relevant_info]['notice_strong_words_total'] += dict_notice_relevant_info[id_dict_noitice_relevant_info]['notice_strong_words_fake_total']

            dict_notice_relevant_info[id_dict_noitice_relevant_info][
                'real_words_strongs_in_notice'] = real_words_strongs_in_notice
            dict_notice_relevant_info[id_dict_noitice_relevant_info][
                'real_words_strongs_in_notice_number_appear'] = real_words_strongs_in_notice_number_appear
            dict_notice_relevant_info[id_dict_noitice_relevant_info][
                'fake_words_strongs_in_notice'] = fake_words_strongs_in_notice
            dict_notice_relevant_info[id_dict_noitice_relevant_info][
                'fake_words_strongs_in_notice_number_appear'] = fake_words_strongs_in_notice_number_appear

    print('Finish update dict of notices with relevant info\n')
    return dict_notice_relevant_info


def load_dict_notices_relevant_info_xlsx(folder_path, file_name, dict_notice_relevant_info):
    print('Starting loading dict of notices relevant info ...')
    file_path = os.path.join(folder_path, file_name)

    df = pd.read_excel(file_path)

    for row_id, row in df.iterrows():
        id_dict_notice = row_id
        id_notice = row['id_notice']
        title_notice = row['title_notice']
        real_words_strongs_in_notice = row['real_words_strongs_in_notice']
        fake_words_strongs_in_notice = row['fake_words_strongs_in_notice']
        notice_strong_words_real_total = row['notice_strong_words_real_total']
        notice_strong_words_fake_total = row['notice_strong_words_fake_total']
        notice_words_total_with_stopwords = row['notice_words_total_with_stopwords']
        notice_strong_words_total = row['notice_strong_words_total']
        classe_notice = row['classe_notice']
        notice_words_total_with_stopwords = row['notice_words_total_with_stopwords']
        notice_words_total_without_stopwords = row['notice_words_total_without_stopwords']

        dict_notice_relevant_info[id_dict_notice] = {}
        dict_notice_relevant_info[id_dict_notice]['id_notice'] = id_notice
        dict_notice_relevant_info[id_dict_notice]['title_notice'] = title_notice
        dict_notice_relevant_info[id_dict_notice]['real_words_strongs_in_notice'] = real_words_strongs_in_notice
        dict_notice_relevant_info[id_dict_notice]['fake_words_strongs_in_notice'] = fake_words_strongs_in_notice
        dict_notice_relevant_info[id_dict_notice]['notice_strong_words_real_total'] = notice_strong_words_real_total
        dict_notice_relevant_info[id_dict_notice]['notice_strong_words_fake_total'] = notice_strong_words_fake_total
        dict_notice_relevant_info[id_dict_notice]['notice_strong_words_total'] = notice_strong_words_total
        dict_notice_relevant_info[id_dict_notice]['classe_notice'] = classe_notice
        dict_notice_relevant_info[id_dict_notice]['notice_words_total_with_stopwords'] = notice_words_total_with_stopwords
        dict_notice_relevant_info[id_dict_notice]['notice_words_total_without_stopwords'] = notice_words_total_without_stopwords

    print('Finish load dict of notices relevant info\n')

    return dict_notice_relevant_info


def save_dict_notices_relevant_info_to_xlsx(file_path, dict_notice_relevant_info, group_name):
    print('Starting save dict of notices relevant info ...')

    df = pd.DataFrame.from_dict(dict_notice_relevant_info, orient='index')

    wb = Workbook()
    ws = wb.active
    ws.title = f'InfoRelDict de Noticias {group_name}'

    headers = list(dict_notice_relevant_info[next(
        iter(dict_notice_relevant_info))].keys())
    for col_num, value in enumerate(headers, start=1):
        ws.cell(row=1, column=col_num, value=value)

    for row_num, (id_notice, row_data) in enumerate(df.iterrows(), start=2):
        for col_num, value in enumerate(row_data, start=1):
            if isinstance(value, list):
                value = '\n'.join(str(value))
            ws.cell(row=row_num, column=col_num, value=value)

    wb.save(file_path)

    print('Finish save dict of notices relevant info\n')


def save_dict_notices_relevant_info_to_csv(file_path, dict_notice_relevant_info, group_name):
    print('Starting save dict of notices relevant info ...')

    pd.set_option('display.max_colwidth', None)

    relevant_columns = [
        'id_notice',
        'real_words_strongs_in_notice',
        'real_words_strongs_in_notice_number_appear',
        'fake_words_strongs_in_notice',
        'fake_words_strongs_in_notice_number_appear',
        'notice_strong_words_real_total',
        'notice_strong_words_fake_total',
        'notice_strong_words_total',
        'classe_notice'
    ]

    df = pd.DataFrame.from_dict(dict_notice_relevant_info, orient='index')[
        relevant_columns]

    df.to_csv(file_path, index=False, encoding='utf-8',
              quoting=csv.QUOTE_NONNUMERIC)

    print('Finish save dict of notices relevant info\n')


def create_dictionary_notices_adapter_to_weka(dict_notice_adapter_to_weka, dict_notice, dict_strong_words, individual_group):
    print('Starting create dict of notices adapter to weka ...')

    if(individual_group == 0):
        for id_dict_notice in dict_notice:
            id_dict_info = len(dict_notice_adapter_to_weka)
            notice_info = dict_notice[id_dict_notice]
            dict_notice_adapter_to_weka[id_dict_info] = {}
            dict_notice_adapter_to_weka[id_dict_info]['id_notice'] = notice_info.get(
                'id_notice')
            dict_notice_adapter_to_weka[id_dict_info]['title_notice'] = notice_info.get(
                'title_notice')
            for id_dict_strong_word in dict_strong_words:
                strong_word_info = dict_strong_words[id_dict_strong_word]
                dict_notice_adapter_to_weka[id_dict_info][strong_word_info.get('word')] = ' '
            dict_notice_adapter_to_weka[id_dict_info]['class_notice'] = notice_info.get('classe_notice')
    else:
        for id_dict_notice_adapter_to_weka in dict_notice_adapter_to_weka:
            for id_dict_strong_word in dict_strong_words:
                strong_word_info = dict_strong_words[id_dict_strong_word]
                dict_notice_adapter_to_weka[id_dict_notice_adapter_to_weka][strong_word_info.get('word')] = ' '

    print('Finish create dict of notices adapter to weka\n')

    return dict_notice_adapter_to_weka

def update_dictionary_notices_adapter_to_weka(dict_notice_adapter_to_weka, dict_notice, dict_strong_word):
    print('Starting update dict of notices adapter to weka ...')

    for id_dict_notice_adapter_to_weka, notice_adapter_to_weka  in dict_notice_adapter_to_weka.items():

        id_notice_adapter_to_weka = notice_adapter_to_weka['id_notice']

        matching_notice = None
        for notice in dict_notice.values():
            if notice['id_notice'] == id_notice_adapter_to_weka:
                matching_notice = notice
                break

        if matching_notice:
            notice_words = matching_notice['notice_content_stemm_without_stopwords']

            for word in notice_words:
                if word in {entry['word'] for entry in dict_strong_word.values()}:
                    dict_notice_adapter_to_weka[id_dict_notice_adapter_to_weka][word] = notice_words.count(word)

    print('Finish update dict of notices adapter to weka\n')
    return dict_notice_adapter_to_weka

def remove_notices_not_appear_strong_words(dict_notice_adapter_to_weka, strong_words_boths_total):
    print('Starting remove dict of notices adapter to weka ...')

    id_notice_to_remove = []

    for id_notice, value in dict_notice_adapter_to_weka.items():
        if isinstance(value, (list, tuple)):
            empty_count = value.count('')

            if empty_count == strong_words_boths_total:
                id_notice_to_remove.append(id_notice)
    
    for id_remove in id_notice_to_remove:
        del dict_notice_adapter_to_weka[id_remove]

    print('Finish remove dict of notices adapter to weka\n')
    return dict_notice_adapter_to_weka

def load_dict_notices_adapter_to_weka_xlsx(dict_notice_adapter_to_weka, file_path):
    print('Starting load dict of notices adapter to weka from xlsx...')

    df = pd.read_excel(file_path)
    
    for index, row in df.iterrows():
        dict_info = {}
        for col_name, value in row.items():
            if col_name == 'id_notice':
                dict_info['id_notice'] = value
            elif col_name == 'title_notice':
                dict_info['title_notice'] = value
            else:
                dict_info[col_name] = value
        dict_notice_adapter_to_weka[index] = dict_info
    
    print('Finish load dict of notices adapter to weka from xlsx\n')
    return dict_notice_adapter_to_weka

def save_dict_notices_adapter_to_weka_to_xlsx(file_path, dict_notice_adapter_to_weka, group_name):
    print('Starting save dict of notices adapter to weka ...')

    df = pd.DataFrame.from_dict(dict_notice_adapter_to_weka, orient='index')

    df['class_notice'] = df.pop('class_notice')

    wb = Workbook()
    ws = wb.active
    ws.title = f'DictToWeka de Noticias {group_name}'

    headers = list(dict_notice_adapter_to_weka[next(
        iter(dict_notice_adapter_to_weka))].keys())
    for col_num, value in enumerate(headers, start=1):
        ws.cell(row=1, column=col_num, value=value)

    for row_num, (id_notice, row_data) in enumerate(df.iterrows(), start=2):
        for col_num, value in enumerate(row_data, start=1):
            if isinstance(value, list):
                value = '\n'.join(str(value))
            ws.cell(row=row_num, column=col_num, value=value)

    wb.save(file_path)

    print('Finish save dict of notices adapter to weka\n')


def save_dict_notices_adapter_to_weka_to_csv(file_path, dict_notice_adapter_to_weka, group_name):
    print('Starting save dict of notices adapter to weka ...')

    df = pd.DataFrame.from_dict(dict_notice_adapter_to_weka, orient='index')

    df['class_notice'] = df.pop('class_notice')

    df = df.drop(columns=['title_notice'])

    df.to_csv(file_path, index_label='id_notice')

    print('Finish save dict of notices adapter to weka\n')